from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .cause_effect import normalise_zone_key


SCHEMA_VERSION = "firepanel-testing-v1"
SESSION_SHEET = "Test Sessions"
RESULT_SHEET = "Test Results"
SESSION_HEADERS = [
    "Session Key",
    "Trigger Zone",
    "Trigger Zone Name",
    "Engineer",
    "Scope Node",
    "Notes",
]
RESULT_HEADERS = [
    "Session Key",
    "Trigger Zone",
    "Stable Key",
    "Expected State",
    "Actual State",
    "Result",
    "Comments",
    "Tested At",
    "Target Node",
    "Target Node Name",
    "Output Group",
    "Output Group Name",
    "Ringing Style",
]


@dataclass(frozen=True, slots=True)
class ImportedTestSession:
    session_key: str
    trigger_zone: str
    engineer: str
    scope_node: int | None
    notes: str
    results: list[tuple[str, str, str, str, str, str | None]]


def export_testing_workbook(
    repository,
    target: str | Path,
    trigger_zones: Iterable[object],
) -> Path:
    destination = Path(target)
    selected = {
        normalise_zone_key(zone)
        for zone in trigger_zones
        if normalise_zone_key(zone)
    }
    if not selected:
        raise ValueError("Select at least one trigger zone.")

    activations = [
        row
        for row in repository.fetch_cause_effect_activations()
        if normalise_zone_key(row["trigger_zone"]) in selected
    ]
    by_zone: dict[str, list] = {zone: [] for zone in selected}
    zone_names: dict[str, str] = {}
    for activation in activations:
        zone = normalise_zone_key(activation["trigger_zone"])
        by_zone.setdefault(zone, []).append(activation)
        name = str(activation["trigger_zone_name"] or "").strip()
        if name:
            zone_names[zone] = name
    for zone in repository.fetch_zones():
        key = normalise_zone_key(zone["number"])
        zone_names.setdefault(key, str(zone["description"] or "").strip())

    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    sessions = workbook.create_sheet(SESSION_SHEET)
    results = workbook.create_sheet(RESULT_SHEET)

    instructions.append(["FirePanel output-group testing workbook"])
    instructions.append(["Schema", SCHEMA_VERSION])
    instructions.append(["Project", repository.name])
    instructions.append([])
    instructions.append(
        [
            "For each trigger zone, set off a fire call and check every listed "
            "output group. Complete Actual State, Result, Comments and Tested At, "
            "then import this workbook back into FirePanel."
        ]
    )
    instructions.append(
        [
            "Do not change Session Key, Trigger Zone, Stable Key, Expected State "
            "or the output-group reference columns."
        ]
    )
    instructions.column_dimensions["A"].width = 105
    instructions.column_dimensions["B"].width = 35
    instructions["A1"].font = Font(size=16, bold=True, color="183153")
    instructions["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    instructions["A6"].alignment = Alignment(wrap_text=True, vertical="top")
    instructions.row_dimensions[5].height = 42
    instructions.row_dimensions[6].height = 38
    instructions.sheet_view.showGridLines = False

    sessions.append(SESSION_HEADERS)
    result_rows = []
    for index, zone in enumerate(sorted(selected, key=_zone_sort_key), start=1):
        session_key = f"zone-{index:04d}-{zone}"
        sessions.append(
            [session_key, zone, zone_names.get(zone, ""), "", "", ""]
        )
        for activation in by_zone.get(zone, []):
            node = int(activation["target_node"])
            group = int(activation["output_group"])
            ringing_style = str(activation["ringing_style"] or "").strip()
            stable_key = (
                f"cause-effect/{zone}/node-{node}/group-{group}/"
                f"style-{ringing_style}"
            )
            result_rows.append(
                [
                    session_key,
                    zone,
                    stable_key,
                    "activated",
                    "not-tested",
                    "not-tested",
                    "",
                    "",
                    node,
                    str(activation["target_node_name"] or ""),
                    group,
                    str(activation["output_group_name"] or ""),
                    ringing_style,
                ]
            )

    results.append(RESULT_HEADERS)
    for row in result_rows:
        results.append(row)

    _style_table_sheet(sessions, len(SESSION_HEADERS), len(selected))
    _style_table_sheet(results, len(RESULT_HEADERS), len(result_rows))
    _set_widths(sessions, [24, 14, 38, 22, 14, 45])
    _set_widths(
        results,
        [24, 14, 56, 17, 20, 16, 42, 22, 14, 34, 16, 42, 18],
    )

    if result_rows:
        last_row = len(result_rows) + 1
        actual_validation = DataValidation(
            type="list",
            formula1='"not-tested,activated,not-activated,unable-to-confirm"',
        )
        result_validation = DataValidation(
            type="list",
            formula1='"not-tested,pass,fail,blocked"',
        )
        results.add_data_validation(actual_validation)
        results.add_data_validation(result_validation)
        actual_validation.add(f"E2:E{last_row}")
        result_validation.add(f"F2:F{last_row}")
        results.conditional_formatting.add(
            f"F2:F{last_row}",
            FormulaRule(
                formula=["$F2=\"pass\""],
                fill=PatternFill("solid", fgColor="D1E7DD"),
            ),
        )
        results.conditional_formatting.add(
            f"F2:F{last_row}",
            FormulaRule(
                formula=["$F2=\"fail\""],
                fill=PatternFill("solid", fgColor="F8D7DA"),
            ),
        )
        results.conditional_formatting.add(
            f"F2:F{last_row}",
            FormulaRule(
                formula=["$F2=\"blocked\""],
                fill=PatternFill("solid", fgColor="FFF3CD"),
            ),
        )

    workbook.properties.title = f"{repository.name} - Output-group testing"
    workbook.properties.subject = SCHEMA_VERSION
    workbook.properties.creator = "FirePanel Commissioning"
    workbook.save(destination)
    return destination


def read_testing_workbook(
    source: str | Path,
) -> list[ImportedTestSession]:
    path = Path(source)
    workbook = load_workbook(path, data_only=False, read_only=False)
    if SESSION_SHEET not in workbook.sheetnames or RESULT_SHEET not in workbook.sheetnames:
        raise ValueError(
            f"The workbook must contain '{SESSION_SHEET}' and '{RESULT_SHEET}'."
        )
    if "Instructions" in workbook.sheetnames:
        schema = str(workbook["Instructions"]["B2"].value or "").strip()
        if schema != SCHEMA_VERSION:
            raise ValueError(
                f"Unsupported testing workbook schema '{schema or 'missing'}'."
            )

    session_rows = _rows_by_header(workbook[SESSION_SHEET], SESSION_HEADERS)
    result_rows = _rows_by_header(workbook[RESULT_SHEET], RESULT_HEADERS)
    sessions: dict[str, dict] = {}
    for row in session_rows:
        key = str(row["Session Key"] or "").strip()
        zone = normalise_zone_key(row["Trigger Zone"])
        if not key or not zone:
            continue
        if key in sessions:
            raise ValueError(f"Duplicate test Session Key '{key}'.")
        scope_value = row["Scope Node"]
        try:
            scope_node = int(scope_value) if scope_value not in (None, "") else None
        except (TypeError, ValueError) as error:
            raise ValueError(f"Invalid Scope Node for session '{key}'.") from error
        sessions[key] = {
            "zone": zone,
            "engineer": str(row["Engineer"] or "").strip(),
            "scope_node": scope_node,
            "notes": str(row["Notes"] or "").strip(),
            "results": [],
        }

    for row in result_rows:
        key = str(row["Session Key"] or "").strip()
        if not key and all(value in (None, "") for value in row.values()):
            continue
        if key not in sessions:
            raise ValueError(f"Test Result references unknown Session Key '{key}'.")
        zone = normalise_zone_key(row["Trigger Zone"])
        if zone != sessions[key]["zone"]:
            raise ValueError(
                f"Trigger Zone for result '{row['Stable Key']}' does not match "
                f"session '{key}'."
            )
        stable_key = str(row["Stable Key"] or "").strip()
        if not stable_key:
            raise ValueError(f"A Test Result in session '{key}' has no Stable Key.")
        expected = str(row["Expected State"] or "").strip().casefold()
        actual = str(row["Actual State"] or "not-tested").strip().casefold()
        result = str(row["Result"] or "not-tested").strip().casefold()
        if actual not in {
            "not-tested",
            "activated",
            "not-activated",
            "unable-to-confirm",
        }:
            raise ValueError(f"Invalid Actual State '{actual}' for '{stable_key}'.")
        if result not in {"not-tested", "pass", "fail", "blocked"}:
            raise ValueError(f"Invalid Result '{result}' for '{stable_key}'.")
        tested_at = _normalise_datetime(row["Tested At"])
        sessions[key]["results"].append(
            (
                stable_key,
                expected or "activated",
                actual,
                result,
                str(row["Comments"] or "").strip(),
                tested_at,
            )
        )

    return [
        ImportedTestSession(
            session_key=key,
            trigger_zone=value["zone"],
            engineer=value["engineer"],
            scope_node=value["scope_node"],
            notes=value["notes"],
            results=value["results"],
        )
        for key, value in sessions.items()
    ]


def _rows_by_header(sheet, required_headers: list[str]) -> list[dict[str, object]]:
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    missing = [header for header in required_headers if header not in headers]
    if missing:
        raise ValueError(
            f"'{sheet.title}' is missing required columns: {', '.join(missing)}"
        )
    positions = {header: headers.index(header) + 1 for header in required_headers}
    return [
        {
            header: sheet.cell(row=row, column=column).value
            for header, column in positions.items()
        }
        for row in range(2, sheet.max_row + 1)
    ]


def _normalise_datetime(value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.isoformat(timespec="seconds")
    return str(value).strip() or None


def _style_table_sheet(sheet, column_count: int, data_count: int) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{_column_letter(column_count)}{max(data_count + 1, 2)}"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="183153")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    sheet.row_dimensions[1].height = 34


def _set_widths(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[_column_letter(index)].width = width


def _column_letter(number: int) -> str:
    letters = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _zone_sort_key(value: str) -> tuple[int, float, str]:
    try:
        return (0, float(value), value)
    except ValueError:
        return (1, 0.0, value.casefold())
