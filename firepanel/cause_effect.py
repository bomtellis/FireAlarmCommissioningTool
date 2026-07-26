from __future__ import annotations

import hashlib
import re
from collections import defaultdict, deque
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


MATRIX_SHEET = "Cause & Effect"
REFERENCE_SHEET = "OutputGroupInfo"
DERIVED_REFERENCE_WARNING = (
    f"'{REFERENCE_SHEET}' was not present; its rows were derived from "
    f"'{MATRIX_SHEET}' by the application."
)
ACTIVATION_CODES = frozenset({"A", "E", "TA", "TE"})
NODE_HEADER = re.compile(
    r"^\s*node\s+(\d+)\s*(?:[-–—]\s*)?(.*?)\s*$",
    re.IGNORECASE,
)


@dataclass(slots=True)
class CauseEffectActivation:
    trigger_zone: str
    trigger_zone_name: str
    target_node: int
    target_node_name: str
    output_group: int
    output_group_name: str
    output_zone_name: str
    activation_code: str
    source_row: int
    source_column: str
    reference_status: str = "matrix_only"


@dataclass(slots=True)
class ReferenceActivation:
    target_node: int
    target_node_name: str
    output_group: int
    output_group_name: str
    trigger_zone: str
    output_zone_name: str
    activation_code: str
    source_row: int


@dataclass(slots=True)
class CauseEffectWorkbook:
    source: Path
    sha256: str
    activations: list[CauseEffectActivation]
    output_groups: list[CauseEffectOutputGroup]
    reference_count: int
    matched_count: int
    matrix_only_count: int
    reference_only_count: int
    activation_codes: list[str]
    reference_source: str = "sheet"
    reference_only: list[ReferenceActivation] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    reference_only_examples: list[str] = field(default_factory=list)


@dataclass(frozen=True, slots=True)
class CauseEffectOutputGroup:
    index: int
    target_node: int
    target_node_name: str
    output_group: int
    output_group_name: str


def read_cause_effect_workbook(path: str | Path) -> CauseEffectWorkbook:
    source = Path(path).resolve()
    if not source.exists():
        raise FileNotFoundError(source)

    workbook = load_workbook(
        source,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        matrix = workbook[_sheet_name(workbook.sheetnames, MATRIX_SHEET)]
        activations, output_groups, matrix_warnings = _read_matrix(matrix)
        reference_name = _optional_sheet_name(
            workbook.sheetnames,
            REFERENCE_SHEET,
        )
        if reference_name is None:
            references = _derive_reference(activations)
            reference_source = "derived"
        else:
            references = _read_reference(workbook[reference_name])
            reference_source = "sheet"
    finally:
        workbook.close()

    matched_count, reference_only, warnings = _check_reference(
        activations,
        references,
    )
    warnings[:0] = matrix_warnings
    if reference_source == "derived":
        for activation in activations:
            if activation.reference_status == "matched":
                activation.reference_status = "derived"
        warnings.insert(0, DERIVED_REFERENCE_WARNING)
    matrix_only_count = sum(
        activation.reference_status == "matrix_only" for activation in activations
    )
    if matrix_only_count:
        warnings.append(
            f"{matrix_only_count} matrix activations were not present in "
            f"{REFERENCE_SHEET}."
        )
    if reference_only:
        warnings.append(
            f"{len(reference_only)} {REFERENCE_SHEET} rows were not present in "
            f"{MATRIX_SHEET}."
        )

    activation_codes = sorted(
        {activation.activation_code for activation in activations}
    )
    return CauseEffectWorkbook(
        source=source,
        sha256=_sha256(source),
        activations=activations,
        output_groups=output_groups,
        reference_count=len(references),
        matched_count=matched_count,
        matrix_only_count=matrix_only_count,
        reference_only_count=len(reference_only),
        activation_codes=activation_codes,
        reference_source=reference_source,
        reference_only=reference_only,
        warnings=warnings,
        reference_only_examples=[
            _reference_label(item) for item in reference_only[:20]
        ],
    )


def _sheet_name(sheet_names: Iterable[str], expected: str) -> str:
    name = _optional_sheet_name(sheet_names, expected)
    if name is not None:
        return name
    raise ValueError(f"Workbook is missing the required '{expected}' sheet.")


def _optional_sheet_name(
    sheet_names: Iterable[str],
    expected: str,
) -> str | None:
    normalised_expected = _normalise_header(expected)
    for name in sheet_names:
        if _normalise_header(name) == normalised_expected:
            return name
    return None


def _read_matrix(
    sheet,
) -> tuple[
    list[CauseEffectActivation],
    list[CauseEffectOutputGroup],
    list[str],
]:
    rows = sheet.iter_rows(values_only=True)
    try:
        names = next(rows)
        node_headers = next(rows)
        group_numbers = next(rows)
    except StopIteration as error:
        raise ValueError(f"'{MATRIX_SHEET}' does not contain its three header rows.") from error

    body_rows = list(rows)
    zone_column, zone_name_column = _matrix_zone_columns(body_rows)
    output_columns, warnings = _output_columns(
        names,
        node_headers,
        group_numbers,
        zone_column + 1,
    )
    if not output_columns:
        raise ValueError(
            f"'{MATRIX_SHEET}' did not contain any node/output-group columns."
        )

    activations: list[CauseEffectActivation] = []
    for source_row, row in enumerate(body_rows, start=4):
        trigger_zone = normalise_zone_key(_value(row, zone_column))
        if not trigger_zone:
            continue
        trigger_zone_name = _text(_value(row, zone_name_column))
        for output in output_columns:
            raw_code = _value(row, output.index)
            if raw_code in (None, ""):
                continue
            code = _activation_code(raw_code)
            if code not in ACTIVATION_CODES:
                continue
            activations.append(
                CauseEffectActivation(
                    trigger_zone=trigger_zone,
                    trigger_zone_name=trigger_zone_name,
                    target_node=output.target_node,
                    target_node_name=output.target_node_name,
                    output_group=output.output_group,
                    output_group_name=output.output_group_name,
                    output_zone_name="",
                    activation_code=code,
                    source_row=source_row,
                    source_column=get_column_letter(output.index + 1),
                )
            )

    if not activations:
        raise ValueError(f"'{MATRIX_SHEET}' did not contain any activations.")
    return activations, output_columns, warnings


def _matrix_zone_columns(
    body_rows: list[tuple[object, ...]],
) -> tuple[int, int]:
    zone_column: int | None = None
    zone_name_column: int | None = None
    for row in body_rows[:8]:
        for index, value in enumerate(row):
            header = _normalise_header(value)
            if header == "newzone":
                zone_column = index
            elif header == "zone" and zone_column is None:
                zone_column = index
            if header.endswith("zonelabel"):
                zone_name_column = index
        if zone_column is not None and zone_name_column is not None:
            break
    if zone_column is None:
        raise ValueError(
            f"'{MATRIX_SHEET}' did not contain a Zone or New Zone column."
        )
    if zone_name_column is None:
        zone_name_column = max(0, zone_column - 2)
    return zone_column, zone_name_column


def _output_columns(
    names: tuple[object, ...],
    node_headers: tuple[object, ...],
    group_numbers: tuple[object, ...],
    start_column: int,
) -> tuple[list[CauseEffectOutputGroup], list[str]]:
    result: list[CauseEffectOutputGroup] = []
    unresolved: list[tuple[int, int, str]] = []
    target_node: int | None = None
    target_node_name = ""
    width = max(len(names), len(node_headers), len(group_numbers))
    for index in range(start_column, width):
        header = _text(_value(node_headers, index))
        if header:
            node = _node_header(header)
            if node is None:
                target_node = None
                target_node_name = ""
            else:
                target_node, target_node_name = node

        output_group = _positive_integer(_value(group_numbers, index))
        output_group_name = _text(_value(names, index))
        if output_group is None or not output_group_name:
            continue
        if target_node is None:
            unresolved.append((index, output_group, output_group_name))
            continue
        result.append(
            CauseEffectOutputGroup(
                index=index,
                target_node=target_node,
                target_node_name=target_node_name,
                output_group=output_group,
                output_group_name=output_group_name,
            )
        )

    inferred: dict[int, list[str]] = defaultdict(list)
    unmapped: list[str] = []
    for index, output_group, output_group_name in unresolved:
        previous = max(
            (output for output in result if output.index < index),
            key=lambda output: output.index,
            default=None,
        )
        following = min(
            (output for output in result if output.index > index),
            key=lambda output: output.index,
            default=None,
        )
        inferred_node: int | None = None
        if (
            previous is not None
            and previous.index == index - 1
            and previous.target_node in inferred
        ):
            inferred_node = previous.target_node
        elif (
            previous is not None
            and following is not None
            and following.target_node == previous.target_node + 2
        ):
            inferred_node = previous.target_node + 1
        elif previous is not None and following is None:
            inferred_node = previous.target_node + 1

        column = get_column_letter(index + 1)
        if inferred_node is None:
            unmapped.append(column)
            continue
        inferred[inferred_node].append(column)
        result.append(
            CauseEffectOutputGroup(
                index=index,
                target_node=inferred_node,
                target_node_name=f"Station {inferred_node}",
                output_group=output_group,
                output_group_name=output_group_name,
            )
        )
    result.sort(key=lambda output: output.index)

    warnings: list[str] = []
    for inferred_node, columns in inferred.items():
        examples = ", ".join(columns[:12])
        if len(columns) > 12:
            examples += ", ..."
        warnings.append(
            f"Inferred node/station {inferred_node} for {len(columns)} "
            f"output-group columns with a blank header ({examples})."
        )
    if unmapped:
        examples = ", ".join(unmapped[:12])
        if len(unmapped) > 12:
            examples += ", ..."
        warnings.append(
            f"Skipped {len(unmapped)} output-group columns whose node/station "
            f"header was blank or unrecognised ({examples})."
        )
    return result, warnings


def _node_header(value: object) -> tuple[int, str] | None:
    header = _text(value)
    match = NODE_HEADER.match(header)
    if match:
        return int(match.group(1)), match.group(2).strip()

    panel = re.match(
        r"^\s*panel\s+0*(\d+)\b\s*(.*?)\s*$",
        header,
        re.IGNORECASE,
    )
    if panel:
        return int(panel.group(1)), panel.group(2).strip()

    station = re.search(r"\bstation\s+0*(\d+)\b", header, re.IGNORECASE)
    if station:
        target_node = int(station.group(1))
        name = re.sub(
            r"\(?\s*station\s+0*\d+\s*\)?",
            "",
            header,
            flags=re.IGNORECASE,
        ).strip(" -()")
        return target_node, name or header

    fa = re.match(r"^\s*FA\s*0*(\d+)\b\s*(.*?)\s*$", header, re.IGNORECASE)
    if fa:
        return int(fa.group(1)), fa.group(2).strip(" -()")
    return None


def _derive_reference(
    activations: list[CauseEffectActivation],
) -> list[ReferenceActivation]:
    zone_names: dict[str, str] = {}
    for activation in activations:
        if activation.trigger_zone_name:
            zone_names.setdefault(
                activation.trigger_zone.casefold(),
                activation.trigger_zone_name,
            )

    return [
        ReferenceActivation(
            target_node=activation.target_node,
            target_node_name=activation.target_node_name,
            output_group=activation.output_group,
            output_group_name=activation.output_group_name,
            trigger_zone=activation.trigger_zone,
            output_zone_name=_derived_output_zone_name(
                activation.output_group_name,
                zone_names,
            ),
            activation_code=activation.activation_code,
            source_row=activation.source_row,
        )
        for activation in activations
    ]


def _derived_output_zone_name(
    output_group_name: str,
    zone_names: dict[str, str],
) -> str:
    if "sounders" not in output_group_name.casefold():
        return ""
    words = output_group_name.split()
    if len(words) < 2:
        return ""
    return zone_names.get(normalise_zone_key(words[1]).casefold(), "")


def _read_reference(sheet) -> list[ReferenceActivation]:
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as error:
        raise ValueError(f"'{REFERENCE_SHEET}' is empty.") from error

    headers = {
        _normalise_header(value): index
        for index, value in enumerate(raw_headers)
        if value not in (None, "")
    }
    required = {
        "nodenumber",
        "nodename",
        "outputgroupnumber",
        "outputgroupname",
        "zonenumber",
        "zonename",
        "activation",
    }
    missing = sorted(required - headers.keys())
    if missing:
        raise ValueError(
            f"'{REFERENCE_SHEET}' is missing required columns: {', '.join(missing)}."
        )

    result: list[ReferenceActivation] = []
    for source_row, row in enumerate(rows, start=2):
        target_node = _positive_integer(_value(row, headers["nodenumber"]))
        output_group = _positive_integer(
            _value(row, headers["outputgroupnumber"])
        )
        trigger_zone = normalise_zone_key(_value(row, headers["zonenumber"]))
        code = _activation_code(_value(row, headers["activation"]))
        if (
            target_node is None
            or output_group is None
            or not trigger_zone
            or not code
        ):
            continue
        result.append(
            ReferenceActivation(
                target_node=target_node,
                target_node_name=_text(_value(row, headers["nodename"])),
                output_group=output_group,
                output_group_name=_text(
                    _value(row, headers["outputgroupname"])
                ),
                trigger_zone=trigger_zone,
                output_zone_name=_text(_value(row, headers["zonename"])),
                activation_code=code,
                source_row=source_row,
            )
        )
    return result


def _check_reference(
    activations: list[CauseEffectActivation],
    references: list[ReferenceActivation],
) -> tuple[int, list[ReferenceActivation], list[str]]:
    by_key: dict[tuple[int, int, str, str], deque[ReferenceActivation]] = (
        defaultdict(deque)
    )
    output_zone_names: dict[tuple[int, int], list[str]] = defaultdict(list)
    for reference in references:
        by_key[_activation_key(reference)].append(reference)
        if (
            reference.output_zone_name
            and reference.output_zone_name.casefold() not in {"n/a", "#value!"}
        ):
            output_zone_names[
                (reference.target_node, reference.output_group)
            ].append(reference.output_zone_name)

    matched = 0
    warnings: list[str] = []
    metadata_mismatches = 0
    for activation in activations:
        candidates = by_key.get(_activation_key(activation))
        if candidates:
            reference = candidates.popleft()
            activation.reference_status = "matched"
            activation.output_zone_name = reference.output_zone_name
            matched += 1
            if (
                _normalise_text(activation.target_node_name)
                != _normalise_text(reference.target_node_name)
                or _normalise_text(activation.output_group_name)
                != _normalise_text(reference.output_group_name)
            ):
                metadata_mismatches += 1
            continue
        fallback_names = output_zone_names.get(
            (activation.target_node, activation.output_group),
            [],
        )
        if fallback_names:
            activation.output_zone_name = fallback_names[0]

    reference_only = [
        reference
        for queue in by_key.values()
        for reference in queue
    ]
    if metadata_mismatches:
        warnings.append(
            f"{metadata_mismatches} matched rows had different node or output-group names."
        )
    return matched, reference_only, warnings


def _activation_key(
    item: CauseEffectActivation | ReferenceActivation,
) -> tuple[int, int, str, str]:
    return (
        item.target_node,
        item.output_group,
        item.trigger_zone.casefold(),
        item.activation_code.casefold(),
    )


def _reference_label(item: ReferenceActivation) -> str:
    return (
        f"Row {item.source_row}: zone {item.trigger_zone}, node "
        f"{item.target_node}, output group {item.output_group}, "
        f"activation {item.activation_code}"
    )


def _positive_integer(value: object) -> int | None:
    if isinstance(value, bool) or value in (None, ""):
        return None
    try:
        parsed = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None
    if parsed != parsed.to_integral_value() or parsed <= 0:
        return None
    return int(parsed)


def normalise_zone_key(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float, Decimal)):
        try:
            number = Decimal(str(value))
            if number == number.to_integral_value():
                return str(int(number))
            return format(number.normalize(), "f")
        except InvalidOperation:
            pass
    return _text(value)


def _activation_code(value: object) -> str:
    return _text(value).upper()


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _normalise_text(value: object) -> str:
    return " ".join(_text(value).casefold().split())


def _normalise_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", _text(value).casefold())


def _value(row: tuple[object, ...], index: int) -> object | None:
    return row[index] if index < len(row) else None


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()
