from __future__ import annotations

from datetime import datetime
from pathlib import Path
import re
import zipfile
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .device_catalog import catalogue_display_name
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .ncf import (
    ConfigurationCauseEffect,
    ConfigurationCauseEffectOutput,
    read_configuration_cause_effect,
)
from .project import ProjectRepository


def export_change_pdf(
    repository: ProjectRepository,
    target: str | Path,
    snapshot_id: int | None = None,
    revision_ids: list[int] | None = None,
) -> Path:
    destination = Path(target)
    changes = repository.fetch_change_details(
        include_all_imports=True,
    )
    history = repository.fetch_project_history()
    revisions = sorted(
        (dict(row) for row in repository.fetch_snapshots()),
        key=lambda row: (str(row["imported_at"]), int(row["id"])),
    )
    if snapshot_id is not None:
        revision_ids = [int(snapshot_id)]
    selected_revision_ids = (
        {int(value) for value in revision_ids}
        if revision_ids is not None
        else {int(row["id"]) for row in revisions}
    )
    revisions = [
        row for row in revisions
        if int(row["id"]) in selected_revision_ids
    ]
    all_revisions = sorted(
        (dict(row) for row in repository.fetch_snapshots()),
        key=lambda row: (str(row["imported_at"]), int(row["id"])),
    )
    changes_by_revision: dict[int, list[dict]] = {
        int(row["id"]): [] for row in revisions
    }
    history_by_revision: dict[int, list[dict]] = {
        int(row["id"]): [] for row in revisions
    }
    for change in changes:
        revision_id = (
            int(change["snapshot_id"])
            if change.get("snapshot_id") is not None
            else _revision_id_at(
                str(change.get("changed_at") or ""),
                all_revisions,
            )
        )
        if revision_id in changes_by_revision:
            changes_by_revision[revision_id].append(change)
    for change in history:
        revision_id = _revision_id_at(
            str(change.get("changed_at") or ""),
            all_revisions,
        )
        if revision_id in history_by_revision:
            history_by_revision[revision_id].append(change)

    selected_changes = sum(
        (changes_by_revision[int(row["id"])] for row in revisions),
        [],
    )
    selected_history = sum(
        (history_by_revision[int(row["id"])] for row in revisions),
        [],
    )
    matrix_import = repository.latest_cause_effect_import()
    styles = getSampleStyleSheet()
    section_style = styles["Heading2"].clone("NodeSection")
    section_style.keepWithNext = True
    section_style.spaceBefore = 5 * mm
    subsection_style = styles["Heading3"].clone("ChangeSubsection")
    subsection_style.keepWithNext = True
    subsection_style.spaceBefore = 2 * mm
    table_text_style = styles["BodyText"].clone("ChangeTableText")
    table_text_style.fontSize = 7
    table_text_style.leading = 8.5

    document = SimpleDocTemplate(
        str(destination),
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=13 * mm,
        bottomMargin=13 * mm,
        title=f"{repository.name} - project tracked changes",
        author="FirePanel",
    )
    story = [
        Paragraph(
            _paragraph_text(
                f"{repository.name} - project tracked changes"
            ),
            styles["Title"],
        ),
        Spacer(1, 4 * mm),
        Paragraph(
            f"Configuration revisions: {len(revisions)} selected of "
            f"{len(all_revisions)}<br/>"
            f"Cause &amp; Effect: "
            f"{matrix_import['source_name'] if matrix_import else 'None'}"
            f"{' - imported ' + matrix_import['imported_at'] if matrix_import else ''}<br/>"
            f"Generated: {datetime.now().isoformat(timespec='minutes')}<br/>"
            "Scope: configuration differences and recorded project history "
            "associated with the selected revisions",
            styles["Normal"],
        ),
        Spacer(1, 5 * mm),
    ]
    if not revisions:
        story.append(
            Paragraph("No revisions were selected.", styles["Heading2"])
        )
    elif not selected_changes and not selected_history:
        story.append(
            Paragraph(
                "No project changes were recorded for the selected revisions.",
                styles["Heading2"],
            )
        )
    else:
        story.extend(
            [
                Paragraph("Total change history", styles["Heading1"]),
                Paragraph(
                    f"<b>{len(selected_changes) + len(selected_history)}</b> "
                    f"recorded changes across {len(revisions)} selected "
                    f"revision{'s' if len(revisions) != 1 else ''}: "
                    f"{len(selected_changes)} configuration/Cause &amp; Effect "
                    f"changes and {len(selected_history)} ongoing project "
                    "history entries.",
                    styles["Normal"],
                ),
                Spacer(1, 2 * mm),
            ]
        )
        revision_numbers = {
            int(row["id"]): index
            for index, row in enumerate(all_revisions, 1)
        }
        for revision in revisions:
            revision_id = int(revision["id"])
            story.append(PageBreak())
            story.append(
                Paragraph(
                    _paragraph_text(
                        f"Revision {revision_numbers[revision_id]} — "
                        f"{revision['source_name']}"
                    ),
                    styles["Heading1"],
                )
            )
            story.append(
                Paragraph(
                    f"Imported {_paragraph_text(revision['imported_at'])}",
                    styles["Normal"],
                )
            )
            _append_revision_changes(
                story,
                changes_by_revision[revision_id],
                history_by_revision[revision_id],
                section_style,
                subsection_style,
                table_text_style,
            )

        story.extend(
            [
                PageBreak(),
                Paragraph("Review and approval", styles["Heading1"]),
                Paragraph(
                    "This report contains configuration and Cause &amp; Effect "
                    "differences together with the recorded project history for "
                    "the selected revisions. It does not by itself approve the fire "
                    "strategy or demonstrate successful cause-and-effect testing.",
                    styles["Normal"],
                ),
                Spacer(1, 12 * mm),
                Paragraph("Reviewed by: ____________________________________", styles["Normal"]),
                Spacer(1, 8 * mm),
                Paragraph("Role: ___________________________________________", styles["Normal"]),
                Spacer(1, 8 * mm),
                Paragraph("Date: ___________________________________________", styles["Normal"]),
            ]
        )
    document.build(
        story,
        onFirstPage=_draw_change_pdf_footer,
        onLaterPages=_draw_change_pdf_footer,
    )
    return destination


def _revision_id_at(
    changed_at: str,
    revisions: list[dict],
) -> int | None:
    if not revisions:
        return None
    revision_id = int(revisions[0]["id"])
    if not changed_at:
        return int(revisions[-1]["id"])
    for revision in revisions:
        if str(revision["imported_at"]) > changed_at:
            break
        revision_id = int(revision["id"])
    return revision_id


def _append_revision_changes(
    story: list,
    changes: list[dict],
    history: list[dict],
    section_style,
    subsection_style,
    table_text_style,
) -> None:
    if not changes and not history:
        story.append(
            Paragraph(
                "No changes were recorded during this revision.",
                subsection_style,
            )
        )
        return
    changes_by_node = _group_changes_by_node(changes)
    history_by_node = _group_changes_by_node(history)
    nodes = sorted(
        {
            node
            for node in set(changes_by_node) | set(history_by_node)
            if node is not None
        }
    )
    for node in nodes:
        node_changes = changes_by_node.get(node, [])
        node_history = history_by_node.get(node, [])
        panel_names = sorted(
            {
                str(row.get("panel") or "").strip()
                for row in node_changes
                if str(row.get("panel") or "").strip()
            }
        )
        node_title = f"Node {node}"
        if panel_names:
            node_title += f" - {', '.join(panel_names)}"
        story.append(Paragraph(_paragraph_text(node_title), section_style))
        if node_changes:
            story.append(
                Paragraph("Configuration changes", subsection_style)
            )
            story.append(
                _configuration_changes_table(
                    node_changes, table_text_style
                )
            )
        if node_history:
            story.append(Paragraph("Project history", subsection_style))
            story.append(
                _project_history_table(node_history, table_text_style)
            )

    project_changes = changes_by_node.get(None, [])
    project_history = history_by_node.get(None, [])
    if project_changes or project_history:
        story.append(Paragraph("Other project changes", section_style))
        if project_changes:
            story.append(
                Paragraph("Configuration changes", subsection_style)
            )
            story.append(
                _configuration_changes_table(
                    project_changes, table_text_style
                )
            )
        if project_history:
            story.append(
                Paragraph("Project history", subsection_style)
            )
            story.append(
                _project_history_table(
                    project_history, table_text_style
                )
            )


def _node_for_change(change: dict) -> int | None:
    node = change.get("node")
    if node not in (None, ""):
        try:
            return int(node)
        except (TypeError, ValueError):
            pass
    match = re.search(
        r"\bnode\s+(\d+)\b",
        str(change.get("stable_key") or ""),
        flags=re.IGNORECASE,
    )
    return int(match.group(1)) if match else None


def _group_changes_by_node(
    changes: list[dict],
) -> dict[int | None, list[dict]]:
    grouped: dict[int | None, list[dict]] = {}
    for change in changes:
        grouped.setdefault(_node_for_change(change), []).append(change)
    return grouped


def _paragraph_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = (
        text.replace("→", "->")
        .replace("—", "-")
        .replace("–", "-")
    )
    return escape(text).replace("\n", "<br/>")


def _draw_change_pdf_footer(canvas, document) -> None:
    canvas.saveState()
    canvas.setFillColor(colors.HexColor("#64748B"))
    canvas.setFont("Helvetica", 7)
    canvas.drawString(
        14 * mm,
        7 * mm,
        "FirePanel project tracked changes",
    )
    canvas.drawRightString(
        landscape(A4)[0] - 14 * mm,
        7 * mm,
        f"Page {document.page}",
    )
    canvas.restoreState()


def _table_cell(value: object, style) -> Paragraph:
    return Paragraph(_paragraph_text(value), style)


def _change_table_style() -> TableStyle:
    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#183153")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 7),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.25,
                colors.HexColor("#CBD5E1"),
            ),
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [colors.white, colors.HexColor("#F8FAFC")],
            ),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
    )


def _configuration_changes_table(changes: list[dict], style) -> Table:
    headers = [
        "Type",
        "Import",
        "Change",
        "Address",
        "Zone",
        "Device text",
        "Device type",
        "Output group",
        "Field",
        "Previous",
        "Current",
    ]
    rows = [headers]
    for change in changes:
        address = ""
        if change.get("loop") is not None:
            address = (
                f"L{change['loop']} / A{change['address']} / "
                f"S{change['sub_address']}"
            )
        output_group = ""
        if change.get("output_group") is not None:
            output_group = str(change["output_group"])
            if str(change.get("output_group_name") or "").strip():
                output_group += f" — {change['output_group_name']}"
            if str(change.get("ringing_style") or "").strip():
                output_group += f" ({change['ringing_style']})"
        rows.append(
            [
                _table_cell(str(change["change_type"]).upper(), style),
                _table_cell(
                    "\n".join(
                        value
                        for value in (
                            str(change.get("changed_at") or ""),
                            str(change.get("source_name") or ""),
                        )
                        if value
                    ),
                    style,
                ),
                _table_cell(change.get("description"), style),
                _table_cell(address or change.get("stable_key"), style),
                _table_cell(change.get("zone"), style),
                _table_cell(change.get("device_text"), style),
                _table_cell(change.get("device_type"), style),
                _table_cell(output_group, style),
                _table_cell(change.get("field"), style),
                _table_cell(change.get("old_value"), style),
                _table_cell(change.get("new_value"), style),
            ]
        )
    table = Table(
        rows,
        repeatRows=1,
        colWidths=[
            15 * mm,
            30 * mm,
            25 * mm,
            24 * mm,
            12 * mm,
            29 * mm,
            23 * mm,
            27 * mm,
            19 * mm,
            32.5 * mm,
            32.5 * mm,
        ],
    )
    table.setStyle(_change_table_style())
    return table


def _project_history_table(history: list[dict], style) -> Table:
    rows = [
        ["Date/time", "Type", "Area", "Record", "Fields", "Description"]
    ]
    rows.extend(
        [
            [
                _table_cell(change.get("changed_at"), style),
                _table_cell(
                    str(change.get("change_type") or "").upper(), style
                ),
                _table_cell(change.get("area"), style),
                _table_cell(change.get("record_key"), style),
                _table_cell(change.get("fields"), style),
                _table_cell(change.get("summary"), style),
            ]
            for change in history
        ]
    )
    table = Table(
        rows,
        repeatRows=1,
        colWidths=[
            34 * mm,
            18 * mm,
            35 * mm,
            45 * mm,
            43 * mm,
            94 * mm,
        ],
    )
    table.setStyle(_change_table_style())
    return table


def export_devices_xlsx(repository: ProjectRepository, target: str | Path) -> Path:
    destination = Path(target)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Devices"
    headers = [
        "Node",
        "Panel",
        "Loop",
        "Address",
        "Sub Address",
        "Zone",
        "Device Text",
        "Device Type",
        "Product Code",
        "Output Group",
        "Test Result",
        "Comments",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="183153")
        cell.alignment = Alignment(horizontal="center")

    for row in repository.fetch_devices():
        sheet.append(
            [
                row["node"],
                row["panel"],
                row["loop"],
                row["address"],
                row["sub_address"],
                row["zone"],
                row["text"],
                catalogue_display_name(row["product_code"], row["observed_type"]),
                row["product_code"],
                row["output_group"],
                "",
                "",
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [9, 34, 9, 10, 13, 10, 38, 20, 14, 15, 15, 40]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    workbook.save(destination)
    return destination


def export_cause_effect_comparison_xlsx(
    repository: ProjectRepository,
    target: str | Path,
) -> Path:
    destination = Path(target)
    snapshots = repository.fetch_snapshots()
    if not snapshots:
        raise ValueError("A configuration must be imported before exporting.")
    snapshot = snapshots[0]

    zone_names = {
        str(row["number"]): str(row["description"] or "").strip()
        for row in repository.fetch_zones()
    }
    imported_activations = repository.fetch_cause_effect_activations()
    imported_outputs = repository.fetch_cause_effect_output_groups()
    for row in imported_activations:
        zone_names.setdefault(
            str(row["trigger_zone"]),
            str(row["trigger_zone_name"] or "").strip(),
        )

    source_path = Path(snapshot["source_path"])
    try:
        configuration = read_configuration_cause_effect(
            source_path,
            zone_names,
        )
    except (OSError, ValueError, zipfile.BadZipFile):
        configuration = ConfigurationCauseEffect(
            source=source_path,
            format_name=source_path.suffix.lstrip(".").upper() or "Configuration",
            warnings=[
                "The source configuration is unavailable or unreadable. "
                "Known project output groups are shown without trigger rules."
            ],
        )
    if not configuration.output_groups:
        configuration.output_groups = [
            ConfigurationCauseEffectOutput(
                target_node=int(row["node"]),
                target_node_name=str(row["panel"] or ""),
                output_group=int(row["output_group"]),
                output_group_name=str(row["group_name"] or ""),
            )
            for row in repository.fetch_output_groups()
        ]

    config_cells, config_raw = _configuration_activation_maps(
        configuration.activations
    )
    imported_cells, imported_details = _imported_activation_maps(
        imported_activations
    )
    all_zones = sorted(
        zone_names.keys()
        | {key[0] for key in config_cells}
        | {key[0] for key in imported_cells},
        key=_zone_sort_key,
    )

    workbook = Workbook()
    configuration_sheet = workbook.active
    configuration_sheet.title = "NCF-SKF Matrix"
    imported_sheet = workbook.create_sheet("Imported Matrix")
    differences_sheet = workbook.create_sheet("Differences")

    config_output_map = {
        (output.target_node, output.output_group): (
            output.target_node_name,
            output.output_group_name,
        )
        for output in configuration.output_groups
    }
    imported_output_map = {
        (int(row["target_node"]), int(row["output_group"])): (
            str(row["target_node_name"] or ""),
            str(row["output_group_name"] or ""),
        )
        for row in imported_outputs
    }
    for row in imported_activations:
        imported_output_map.setdefault(
            (int(row["target_node"]), int(row["output_group"])),
            (
                str(row["target_node_name"] or ""),
                str(row["output_group_name"] or ""),
            ),
        )

    _write_cause_effect_matrix_sheet(
        configuration_sheet,
        title=f"{configuration.format_name} Cause & Effect matrix",
        source_label=str(source_path),
        outputs=config_output_map,
        zones=all_zones,
        zone_names=zone_names,
        cells=config_cells,
        warnings=configuration.warnings,
    )
    imported = repository.latest_cause_effect_import()
    _write_cause_effect_matrix_sheet(
        imported_sheet,
        title="Imported Cause & Effect matrix",
        source_label=(
            str(imported["source_path"])
            if imported is not None
            else "No Cause & Effect workbook imported"
        ),
        outputs=imported_output_map,
        zones=all_zones,
        zone_names=zone_names,
        cells=imported_cells,
        warnings=(
            []
            if imported is not None
            else ["No imported matrix is available for comparison."]
        ),
    )
    _write_difference_sheet(
        differences_sheet,
        config_cells=config_cells,
        config_raw=config_raw,
        imported_cells=imported_cells,
        imported_details=imported_details,
        outputs={**config_output_map, **imported_output_map},
        zone_names=zone_names,
    )

    workbook.properties.title = (
        f"{repository.name} - Cause & Effect comparison"
    )
    workbook.properties.subject = (
        "Configuration, imported matrix and semantic differences"
    )
    workbook.properties.creator = "FirePanel Commissioning"
    workbook.save(destination)
    return destination


def _configuration_activation_maps(
    activations,
) -> tuple[
    dict[tuple[str, int, int], set[str]],
    dict[tuple[str, int, int], set[str]],
]:
    cells: dict[tuple[str, int, int], set[str]] = {}
    raw: dict[tuple[str, int, int], set[str]] = {}
    for activation in activations:
        key = (
            str(activation.trigger_zone),
            int(activation.target_node),
            int(activation.output_group),
        )
        cells.setdefault(key, set()).add(str(activation.ringing_style))
        raw.setdefault(key, set()).add(str(activation.ringing_style_name))
    return cells, raw


def _imported_activation_maps(
    activations,
) -> tuple[
    dict[tuple[str, int, int], set[str]],
    dict[tuple[str, int, int], dict[str, set[str]]],
]:
    cells: dict[tuple[str, int, int], set[str]] = {}
    details: dict[tuple[str, int, int], dict[str, set[str]]] = {}
    for activation in activations:
        key = (
            str(activation["trigger_zone"]),
            int(activation["target_node"]),
            int(activation["output_group"]),
        )
        cells.setdefault(key, set()).add(str(activation["ringing_style"]))
        detail = details.setdefault(
            key,
            {"reference": set(), "comments": set()},
        )
        detail["reference"].add(str(activation["reference_status"]))
        comments = str(activation["comments"] or "").strip()
        if comments:
            detail["comments"].add(comments)
    return cells, details


def _write_cause_effect_matrix_sheet(
    sheet,
    *,
    title: str,
    source_label: str,
    outputs: dict[tuple[int, int], tuple[str, str]],
    zones: list[str],
    zone_names: dict[str, str],
    cells: dict[tuple[str, int, int], set[str]],
    warnings: list[str],
) -> None:
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = title
    sheet["A1"].font = Font(size=16, bold=True, color="183153")
    sheet["A2"] = f"Source: {source_label}"
    sheet["A2"].font = Font(color="475569", italic=True)
    row = 3
    for warning in warnings:
        sheet.cell(row=row, column=1, value=f"Note: {warning}")
        sheet.cell(row=row, column=1).font = Font(color="B45309", italic=True)
        row += 1
    header_row = row + 1

    output_keys = sorted(outputs)
    headers = ["Trigger Zone", "Zone Name"]
    for node, group in output_keys:
        node_name, group_name = outputs[(node, group)]
        header = f"Node {node}"
        if node_name:
            header += f" - {node_name}"
        header += f"\nOutput group {group}"
        if group_name:
            header += f"\n{group_name}"
        headers.append(header)
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="183153")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    sheet.row_dimensions[header_row].height = 66

    thin = Side(style="thin", color="D6DEE8")
    for row_index, zone in enumerate(zones, start=header_row + 1):
        sheet.cell(row=row_index, column=1, value=zone)
        sheet.cell(
            row=row_index,
            column=2,
            value=zone_names.get(zone, ""),
        )
        for column, (node, group) in enumerate(output_keys, start=3):
            values = cells.get((zone, node, group))
            if not values:
                continue
            text = " / ".join(sorted(values))
            cell = sheet.cell(row=row_index, column=column, value=text)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                "solid",
                fgColor=_ringing_style_fill(text),
            )
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin,
            )

    sheet.freeze_panes = f"C{header_row + 1}"
    sheet.auto_filter.ref = (
        f"A{header_row}:B{header_row + max(len(zones), 1)}"
    )
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 38
    for column in range(3, len(output_keys) + 3):
        sheet.column_dimensions[get_column_letter(column)].width = 21
    sheet.sheet_view.zoomScale = 70


def _write_difference_sheet(
    sheet,
    *,
    config_cells: dict[tuple[str, int, int], set[str]],
    config_raw: dict[tuple[str, int, int], set[str]],
    imported_cells: dict[tuple[str, int, int], set[str]],
    imported_details: dict[tuple[str, int, int], dict[str, set[str]]],
    outputs: dict[tuple[int, int], tuple[str, str]],
    zone_names: dict[str, str],
) -> None:
    all_keys = sorted(
        config_cells.keys() | imported_cells.keys(),
        key=lambda key: (_zone_sort_key(key[0]), key[1], key[2]),
    )
    matches = 0
    difference_rows = []
    for zone, node, group in all_keys:
        config_styles = config_cells.get((zone, node, group), set())
        imported_styles = imported_cells.get((zone, node, group), set())
        if config_styles == imported_styles:
            matches += 1
            continue
        if config_styles and imported_styles:
            status = "Different ringing style"
        elif config_styles:
            status = "Configuration only"
        else:
            status = "Imported only"
        node_name, group_name = outputs.get((node, group), ("", ""))
        detail = imported_details.get(
            (zone, node, group),
            {"reference": set(), "comments": set()},
        )
        difference_rows.append(
            [
                status,
                zone,
                zone_names.get(zone, ""),
                node,
                node_name,
                group,
                group_name,
                " / ".join(sorted(config_styles)),
                " / ".join(sorted(imported_styles)),
                " / ".join(
                    sorted(config_raw.get((zone, node, group), set()))
                ),
                " / ".join(sorted(detail["reference"])),
                " | ".join(sorted(detail["comments"])),
            ]
        )

    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "Cause & Effect differences"
    sheet["A1"].font = Font(size=16, bold=True, color="183153")
    sheet["A2"] = (
        f"{matches:,} matched links; "
        f"{len(difference_rows):,} differences."
    )
    sheet["A2"].font = Font(color="475569")
    headers = [
        "Status",
        "Trigger Zone",
        "Zone Name",
        "Node",
        "Node Name",
        "Output Group",
        "Output Group Name",
        "NCF/SKF Style",
        "Imported Style",
        "NCF/SKF Ringing Style",
        "Imported Reference Check",
        "Comments",
    ]
    header_row = 4
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="183153")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    for row_index, values in enumerate(difference_rows, start=header_row + 1):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row=row_index, column=1).fill = PatternFill(
            "solid",
            fgColor={
                "Different ringing style": "FFF3CD",
                "Configuration only": "D1E7DD",
                "Imported only": "F8D7DA",
            }[values[0]],
        )

    last_row = header_row + max(len(difference_rows), 1)
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A{header_row}:L{last_row}"
    widths = [24, 14, 36, 10, 32, 15, 38, 18, 18, 28, 26, 42]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.sheet_view.zoomScale = 85


def _ringing_style_fill(value: str) -> str:
    values = {part.strip().upper() for part in value.split("/")}
    if values == {"E"}:
        return "D1E7DD"
    if "TA" in values:
        return "FFF3CD"
    if "TE" in values:
        return "FCE5CD"
    if "A" in values:
        return "DDEBF7"
    return "E9EFF7"


def _zone_sort_key(value: object) -> tuple[int, float, str]:
    text = str(value).strip()
    try:
        return (0, float(text), text)
    except ValueError:
        return (1, 0.0, text.casefold())
