from __future__ import annotations

import json
import sqlite3
import struct
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from firepanel.cause_effect import (
    DERIVED_REFERENCE_WARNING,
    read_cause_effect_workbook,
)
from firepanel.project import ProjectRepository, zone_shape_key
from firepanel.rules import generate_door_rules
from firepanel.testing_workbook import (
    RESULT_HEADERS,
    ZONE_HEADERS,
    export_testing_workbook,
    read_testing_workbook,
)


def test_custom_rules_can_be_updated_and_removed_but_suggestions_cannot(
    tmp_path: Path,
) -> None:
    project_path = tmp_path / "custom-rules.fcp"
    sqlite3.connect(project_path).close()
    repository = ProjectRepository(project_path)
    with repository.connection() as connection:
        connection.executemany(
            "INSERT INTO zones(number, description) VALUES (?, ?)",
            [(10, "Ward"), (11, "Corridor")],
        )

    rule_id = repository.add_rule(
        "Original rule",
        10,
        "exact",
        11,
        None,
        20,
        "ACTIVATE OUTPUT",
        "Original notes",
    )
    repository.update_custom_rule(
        rule_id,
        "Updated rule",
        11,
        "adjacent",
        10,
        2,
        30,
        "ALERT / intermittent",
        "Updated notes",
    )

    updated = repository.fetch_rule(rule_id)
    assert updated["name"] == "Updated rule"
    assert updated["trigger_zone"] == 11
    assert updated["relation"] == "adjacent"
    assert updated["target_zone"] == 10
    assert updated["target_node"] == 2
    assert updated["output_group"] == 30
    assert updated["action"] == "ALERT / intermittent"
    assert updated["notes"] == "Updated notes"

    with repository.connection() as connection:
        connection.execute(
            "UPDATE cause_effect_rules SET source = ? WHERE id = ?",
            ("HTM 05-03 Figure 2", rule_id),
        )
    with pytest.raises(ValueError, match="Only custom"):
        repository.update_custom_rule(
            rule_id,
            "Blocked",
            10,
            "exact",
            11,
            None,
            None,
            "ALERT / intermittent",
        )
    with pytest.raises(ValueError, match="Only custom"):
        repository.delete_custom_rule(rule_id)
    other_id = repository.add_rule(
        "Second custom rule",
        10,
        "exact",
        11,
        None,
        None,
        "ALERT / intermittent",
    )
    with pytest.raises(ValueError, match="Only custom"):
        repository.delete_custom_rules([rule_id, other_id])
    assert repository.fetch_rule(other_id) is not None

    repository.delete_rules([rule_id, other_id])
    assert repository.fetch_rule(rule_id) is None
    assert repository.fetch_rule(other_id) is None


def test_project_history_keeps_an_append_only_log_of_project_edits(
    tmp_path: Path,
) -> None:
    project_path = tmp_path / "project-history.fcp"
    sqlite3.connect(project_path).close()
    repository = ProjectRepository(project_path)

    repository.set_metadata("project_name", "History test")
    floor_id = repository.add_floor("Ground", 0)
    with repository.connection() as connection:
        connection.execute(
            "INSERT INTO zones(number, description) VALUES (10, 'Ward')"
        )
    repository.assign_zone_geometry(
        10,
        floor_id,
        [(0, 0), (10, 0), (10, 10), (0, 0)],
        "USER_DRAWN",
    )
    geometry_id = repository.fetch_zone_geometry()[0]["id"]
    repository.update_zone_geometry(
        geometry_id,
        [(2, 2), (12, 2), (12, 12), (2, 2)],
    )
    history_before_delete = repository.fetch_project_history()
    repository.remove_zone_geometry(geometry_id)
    history = repository.fetch_project_history()

    assert len(history) >= len(history_before_delete) + 1
    assert any(
        row["area"] == "Zone drawings"
        and row["change_type"] == "removed"
        for row in history
    )
    assert any(
        row["area"] == "Zone drawings"
        and row["change_type"] == "modified"
        and "Geometry Json" in row["summary"]
        for row in history
    )
    assert any(
        row["area"] == "Project details"
        and row["change_type"] == "added"
        for row in history
    )
    history_count = len(history)
    reopened = ProjectRepository(project_path)
    assert len(reopened.fetch_project_history()) == history_count


def test_configuration_change_details_include_full_device_context(
    tmp_path: Path,
) -> None:
    project_path = tmp_path / "change-details.fcp"
    sqlite3.connect(project_path).close()
    repository = ProjectRepository(project_path)
    with repository.connection() as connection:
        old_snapshot = connection.execute(
            """
            INSERT INTO snapshots(
                imported_at, source_name, source_path, sha256,
                versions_json, warnings_json
            ) VALUES ('old', 'old.skf', 'old.skf', 'old', '[]', '[]')
            """
        ).lastrowid
        new_snapshot = connection.execute(
            """
            INSERT INTO snapshots(
                imported_at, source_name, source_path, sha256,
                versions_json, warnings_json
            ) VALUES ('new', 'new.skf', 'new.skf', 'new', '[]', '[]')
            """
        ).lastrowid
        for snapshot_id, zone in (
            (old_snapshot, 10),
            (new_snapshot, 11),
        ):
            connection.execute(
                """
                INSERT INTO devices(
                    snapshot_id, stable_key, node, panel, loop, address,
                    sub_address, zone, text, product_code, observed_type,
                    output_group, output_group_name, ringing_style,
                    record_offset
                ) VALUES (
                    ?, '7/2/15/3', 7, 'Panel 7', 2, 15, 3, ?,
                    'Ward beacon', 27, 'Visual Alarm Device',
                    4, 'Ward beacons', 'Alert', 100
                )
                """,
                (snapshot_id, zone),
            )
        connection.execute(
            """
            INSERT INTO changes(
                snapshot_id, entity, stable_key, change_type,
                field, old_value, new_value
            ) VALUES (?, 'device', '7/2/15/3', 'modified', 'zone', '10', '11')
            """,
            (new_snapshot,),
        )
        connection.execute(
            """
            INSERT INTO changes(
                snapshot_id, entity, stable_key, change_type,
                field, old_value, new_value
            ) VALUES (
                ?, 'device', '7/2/15/3', 'added',
                NULL, NULL, 'Initial device'
            )
            """,
            (old_snapshot,),
        )
        connection.execute(
            """
            INSERT INTO changes(
                snapshot_id, entity, stable_key, change_type,
                field, old_value, new_value
            ) VALUES (
                ?, 'device', '7/2/15/3', 'modified',
                'Location text', 'Old ward beacon', 'Ward beacon'
            )
            """,
            (new_snapshot,),
        )

    detail = next(
        row
        for row in repository.fetch_change_details()
        if row["entity"] == "device"
    )
    assert detail["description"] == "Zone changed"
    assert (
        detail["node"],
        detail["zone"],
        detail["loop"],
        detail["address"],
        detail["sub_address"],
    ) == (7, 11, 2, 15, 3)
    assert detail["device_text"] == "Ward beacon"
    assert detail["device_type"] == "Visual Alarm Device"
    assert detail["output_group"] == 4
    assert detail["output_group_name"] == "Ward beacons"
    assert detail["ringing_style"] == "Alert"
    device_text_detail = next(
        row
        for row in repository.fetch_change_details()
        if row["raw_field"] == "Location text"
    )
    assert device_text_detail["field"] == "Device text"
    assert device_text_detail["description"] == "Device text changed"
    total_history = repository.fetch_change_details(
        include_all_imports=True
    )
    assert len(total_history) == 3
    assert [row["source_name"] for row in total_history] == [
        "new.skf",
        "new.skf",
        "old.skf",
    ]


def test_configuration_changes_include_output_groups_by_node(
    tmp_path: Path,
) -> None:
    project_path = tmp_path / "output-group-changes.fcp"
    sqlite3.connect(project_path).close()
    repository = ProjectRepository(project_path)

    def add_group_line(
        connection,
        snapshot_id,
        source_row,
        group,
        name,
        style_number,
        style,
        style_name,
        zone_from,
        zone_to,
    ):
        connection.execute(
            """
            INSERT INTO configuration_output_group_lines(
                snapshot_id, source_row, target_node, target_node_name,
                output_group, output_group_name, operation,
                output_style_number, ringing_style, ringing_style_name,
                zone_from, zone_to, zone_qualifiers
            ) VALUES (?, ?, 7, 'Panel Seven', ?, ?, 0, ?, ?, ?, ?, ?, 0)
            """,
            (
                snapshot_id,
                source_row,
                group,
                name,
                style_number,
                style,
                style_name,
                zone_from,
                zone_to,
            ),
        )

    with repository.connection() as connection:
        old_snapshot = connection.execute(
            """
            INSERT INTO snapshots(
                imported_at, source_name, source_path, sha256,
                versions_json, warnings_json
            ) VALUES ('old', 'old.skf', 'old.skf', 'og-old', '[]', '[]')
            """
        ).lastrowid
        new_snapshot = connection.execute(
            """
            INSERT INTO snapshots(
                imported_at, source_name, source_path, sha256,
                versions_json, warnings_json
            ) VALUES ('new', 'new.skf', 'new.skf', 'og-new', '[]', '[]')
            """
        ).lastrowid
        add_group_line(
            connection, old_snapshot, 1, 10, "Ward sounders",
            0, "E", "Evacuate", 10, 10,
        )
        add_group_line(
            connection, old_snapshot, 2, 20, "Removed doors",
            0, "E", "Evacuate", 20, 20,
        )
        add_group_line(
            connection, new_snapshot, 1, 10, "Ward alert sounders",
            2, "A", "Alert", 10, 12,
        )
        add_group_line(
            connection, new_snapshot, 3, 30, "New beacons",
            0, "E", "Evacuate", 30, 30,
        )
        changes = repository._calculate_changes(
            connection, old_snapshot, new_snapshot
        )
        connection.executemany(
            """
            INSERT INTO changes(
                snapshot_id, entity, stable_key, change_type,
                field, old_value, new_value
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    new_snapshot,
                    change.entity,
                    change.stable_key,
                    change.change_type,
                    change.field,
                    change.old_value,
                    change.new_value,
                )
                for change in changes
            ],
        )

    output_changes = [
        change for change in changes
        if change.entity == "output_group"
    ]
    assert {
        (change.stable_key, change.change_type, change.field)
        for change in output_changes
    } == {
        ("node 7/group 10", "modified", "output_group_name"),
        ("node 7/group 10", "modified", "zone_triggers"),
        ("node 7/group 10", "modified", "ringing_styles"),
        ("node 7/group 20", "removed", None),
        ("node 7/group 30", "added", None),
    }
    details = [
        row
        for row in repository.fetch_change_details(new_snapshot)
        if row["entity"] == "output_group"
    ]
    assert {row["node"] for row in details} == {7}
    assert {row["output_group"] for row in details} == {10, 20, 30}
    assert {
        row["panel"] for row in details
    } == {"Panel Seven"}
    assert {
        row["field"] for row in details
        if row["change_type"] == "modified"
    } == {
        "Output group name",
        "Zone trigger extent",
        "Ringing styles",
    }
    before_backfill = len(repository.fetch_changes(new_snapshot))
    repository._backfill_configuration_output_group_changes()
    assert len(repository.fetch_changes(new_snapshot)) == before_backfill


def _workbook(path: Path) -> Path:
    workbook = Workbook()
    matrix = workbook.active
    matrix.title = "Cause & Effect"
    matrix["F1"] = "ZONE 1 SOUNDERS"
    matrix["G1"] = "FIRE DOORS"
    matrix["H1"] = "UNUSED SPARE OUTPUT"
    matrix["F2"] = "Node 10 Panel A"
    matrix["F3"] = 1
    matrix["G3"] = 50
    matrix["H3"] = 99
    matrix["C4"] = "32 CHARACTER ZONE LABEL"
    matrix["D4"] = "Old Zone"
    matrix["E4"] = "New Zone"
    matrix["C7"] = "Ward 1"
    matrix["E7"] = 1
    matrix["F7"] = "E"
    matrix["G7"] = "TA"
    matrix["C8"] = "Ward 1 detector 10"
    matrix["E8"] = 1.11
    matrix["F8"] = "TE"

    reference = workbook.create_sheet("OutputGroupInfo")
    reference.append(
        [
            "Node Number",
            "Node Name",
            "Output Group Number",
            "Output Group Name",
            "Zone Number",
            "Zone Name",
            "Activation",
        ]
    )
    reference.append([10, "Panel A", 1, "ZONE 1 SOUNDERS", 1, "Ward 1", "E"])
    reference.append([10, "Panel A", 50, "FIRE DOORS", 1, "N/A", "TA"])
    reference.append([10, "Panel A", 1, "ZONE 1 SOUNDERS", 99, "Ward 99", "E"])
    workbook.save(path)
    return path


def test_reads_matrix_and_checks_output_group_info(tmp_path: Path) -> None:
    parsed = read_cause_effect_workbook(_workbook(tmp_path / "matrix.xlsx"))

    assert len(parsed.activations) == 3
    assert [
        (output.target_node, output.output_group)
        for output in parsed.output_groups
    ] == [(10, 1), (10, 50), (10, 99)]
    assert parsed.reference_count == 3
    assert parsed.matched_count == 2
    assert parsed.matrix_only_count == 1
    assert parsed.reference_only_count == 1
    assert len(parsed.reference_only) == 1
    assert parsed.activation_codes == ["E", "TA", "TE"]

    timed = next(
        activation
        for activation in parsed.activations
        if activation.output_group == 50
    )
    assert timed.trigger_zone == "1"
    assert timed.target_node == 10
    assert timed.target_node_name == "Panel A"
    assert timed.activation_code == "TA"
    assert timed.reference_status == "matched"

    matrix_only = next(
        activation
        for activation in parsed.activations
        if activation.activation_code == "TE"
    )
    assert matrix_only.trigger_zone == "1.11"
    assert matrix_only.reference_status == "matrix_only"


def test_derives_output_group_info_when_reference_sheet_is_missing(
    tmp_path: Path,
) -> None:
    path = _workbook(tmp_path / "matrix-without-reference.xlsx")
    workbook = load_workbook(path)
    del workbook["OutputGroupInfo"]
    workbook.save(path)
    workbook.close()

    parsed = read_cause_effect_workbook(path)

    assert parsed.reference_source == "derived"
    assert parsed.reference_count == 3
    assert parsed.matched_count == 3
    assert parsed.matrix_only_count == 0
    assert parsed.reference_only_count == 0
    assert {row.reference_status for row in parsed.activations} == {"derived"}
    assert DERIVED_REFERENCE_WARNING in parsed.warnings
    sounders = next(
        activation
        for activation in parsed.activations
        if activation.output_group == 1
        and activation.trigger_zone == "1"
    )
    assert sounders.output_zone_name == "Ward 1"


def test_reads_legacy_matrix_layout_without_output_group_info(
    tmp_path: Path,
) -> None:
    path = tmp_path / "legacy-matrix.xlsx"
    workbook = Workbook()
    matrix = workbook.active
    matrix.title = "Cause & Effect"
    matrix["D1"] = "GROUND FLOOR SOUNDERS"
    matrix["E1"] = "GROUND FLOOR DOORS"
    matrix.merge_cells("D2:E2")
    matrix["D2"] = "Panel 01 Tower Block"
    matrix["D3"] = 2
    matrix["E3"] = 7
    matrix["A4"] = "Stn No"
    matrix["B4"] = "40 CHARACTER ZONE LABEL"
    matrix["C4"] = "Zone"
    matrix["A7"] = 1
    matrix["B7"] = "Ground floor ward"
    matrix["C7"] = "1A"
    matrix["D7"] = "E"
    matrix["E7"] = "A"
    matrix["F2"] = "Stn No"
    matrix["F3"] = "O/P Group"
    matrix["G1"] = "REMOTE SOUNDERS"
    matrix["G2"] = "FA2 (Station 2)"
    matrix["G3"] = 1
    matrix["G7"] = "TA"
    workbook.save(path)
    workbook.close()

    parsed = read_cause_effect_workbook(path)

    assert parsed.reference_source == "derived"
    assert [
        (
            output.target_node,
            output.target_node_name,
            output.output_group,
        )
        for output in parsed.output_groups
    ] == [
        (1, "Tower Block", 2),
        (1, "Tower Block", 7),
        (2, "FA2", 1),
    ]
    assert {
        (
            activation.trigger_zone,
            activation.trigger_zone_name,
            activation.target_node,
            activation.output_group,
            activation.activation_code,
        )
        for activation in parsed.activations
    } == {
        ("1A", "Ground floor ward", 1, 2, "E"),
        ("1A", "Ground floor ward", 1, 7, "A"),
        ("1A", "Ground floor ward", 2, 1, "TA"),
    }
    assert parsed.matched_count == 3
    assert parsed.matrix_only_count == 0


def test_project_can_start_with_workbook_then_import_configuration(
    tmp_path: Path,
) -> None:
    workbook_path = _workbook(tmp_path / "initial-cause-effect.xlsx")
    workbook = load_workbook(workbook_path)
    del workbook["OutputGroupInfo"]
    workbook.save(workbook_path)
    workbook.close()

    repository = ProjectRepository.create_from_source(
        tmp_path / "workbook-first.fcp",
        "Workbook first",
        workbook_path,
    )
    assert repository.latest_snapshot_id() is None
    assert repository.latest_cause_effect_import() is not None
    assert len(repository.fetch_cause_effect_activations()) == 3

    configuration_path = tmp_path / "site-after-workbook.ncf"
    site = bytearray(336)
    panel_name = b"Panel A"
    site[112 + 8] = 0x12
    site[112 + 9] = len(panel_name)
    site[112 + 10 : 112 + 10 + len(panel_name)] = panel_name
    site[112 + 44] = 5
    struct.pack_into("<i", site, 224 + 8, 3)
    struct.pack_into("<i", site, 224 + 12, 7)
    zone_name = b"Legacy zone"
    site[224 + 16] = len(zone_name)
    site[224 + 17 : 224 + 17 + len(zone_name)] = zone_name

    point = bytearray(224)
    struct.pack_into("<i", point, 8, 2)
    struct.pack_into("<i", point, 12, 45)
    point[16] = 1
    point[17] = 4
    point[18] = 2
    point[20] = len(b"Legacy detector")
    point[21 : 21 + len(b"Legacy detector")] = b"Legacy detector"
    struct.pack_into("<i", point, 48, 7)
    with zipfile.ZipFile(
        configuration_path,
        "w",
        compression=zipfile.ZIP_DEFLATED,
    ) as archive:
        archive.writestr("SITE", site)
        archive.writestr("Panel A.pcf", point)

    snapshot_id, changes = repository.import_configuration(configuration_path)

    assert snapshot_id == 1
    assert changes
    assert len(repository.fetch_panels()) == 1
    assert len(repository.fetch_devices()) == 1
    assert len(repository.fetch_cause_effect_activations()) == 3


def test_project_import_persists_activations_and_comments(tmp_path: Path) -> None:
    workbook_path = _workbook(tmp_path / "matrix.xlsx")
    project_path = tmp_path / "project.fcp"
    sqlite3.connect(project_path).close()
    repository = ProjectRepository(project_path)

    import_id, parsed = repository.import_cause_effect(workbook_path)
    assert parsed.matched_count == 2
    initial_changes = repository.fetch_cause_effect_changes(import_id)
    assert len(initial_changes) == 1
    assert initial_changes[0]["entity"] == "cause_effect_matrix"
    assert initial_changes[0]["change_type"] == "initial"
    assert initial_changes[0]["new_value"] == "3"
    assert repository.latest_cause_effect_import()["id"] == import_id
    assert len(repository.fetch_cause_effect_activations()) == 3
    assert [
        (row["target_node"], row["output_group"])
        for row in repository.fetch_cause_effect_output_groups()
    ] == [(10, 1), (10, 50), (10, 99)]
    assert len(repository.fetch_cause_effect_reference_only()) == 1
    assert len(repository.fetch_cause_effect_activations("1")) == 2
    assert len(repository.fetch_cause_effect_activations(1.11)) == 1

    activation = repository.fetch_cause_effect_activations("1")[0]
    repository.update_cause_effect_comment(
        activation["id"],
        "Witness with maternity team",
    )
    updated = repository.fetch_cause_effect_activations("1")[0]
    assert updated["comments"] == "Witness with maternity team"

    duplicate_id, _ = repository.import_cause_effect(workbook_path)
    assert duplicate_id == import_id
    assert len(repository.fetch_cause_effect_activations()) == 3

    revised_path = tmp_path / "matrix-revised.xlsx"
    revised = load_workbook(workbook_path)
    revised["Cause & Effect"]["A20"] = "Revision 2"
    revised.save(revised_path)
    revised.close()
    revised_id, _ = repository.import_cause_effect(revised_path)
    assert revised_id != import_id
    assert repository.fetch_cause_effect_activations("1")[0]["comments"] == (
        "Witness with maternity team"
    )
    assert [
        row["trigger_zone"] for row in repository.fetch_cause_effect_trigger_zones()
    ] == ["1", "1.11"]
    assert repository.fetch_cause_effect_changes(revised_id) == []


def test_testing_workbook_round_trips_selected_zone_output_groups(
    tmp_path: Path,
) -> None:
    project_path = tmp_path / "project.fcp"
    sqlite3.connect(project_path).close()
    repository = ProjectRepository(project_path)
    repository.import_cause_effect(_workbook(tmp_path / "matrix.xlsx"))

    workbook_path = tmp_path / "testing.xlsx"
    export_testing_workbook(repository, workbook_path, ["1"])
    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames == [
        "Instructions",
        "Zone List",
        "Test Sessions",
        "Test Results",
    ]
    zone_list = workbook["Zone List"]
    assert [cell.value for cell in zone_list[1]] == ZONE_HEADERS
    assert zone_list.freeze_panes == "A2"
    assert {
        (
            str(zone_list.cell(row=row, column=1).value),
            zone_list.cell(row=row, column=3).value,
        )
        for row in range(2, zone_list.max_row + 1)
    } == {
        ("1", "Yes"),
        ("1.11", "No"),
    }
    results = workbook["Test Results"]
    assert [cell.value for cell in results[1]] == RESULT_HEADERS
    assert results.freeze_panes == "A2"
    assert results.max_row == 3
    assert {results.cell(row=row, column=2).value for row in (2, 3)} == {"1"}
    assert {
        results.cell(row=row, column=11).value for row in (2, 3)
    } == {1, 50}

    workbook["Test Sessions"]["D2"] = "Commissioning Engineer"
    results["E2"] = "activated"
    results["F2"] = "pass"
    results["G2"] = "Witnessed at panel"
    results["H2"] = "2026-07-24T10:30:00+00:00"
    results["E3"] = "not-activated"
    results["F3"] = "fail"
    workbook.save(workbook_path)

    imported = read_testing_workbook(workbook_path)
    assert len(imported) == 1
    assert imported[0].trigger_zone == "1"
    assert len(imported[0].results) == 2
    assert repository.import_test_sessions(imported) == (1, 2)
    stored = repository.fetch_test_results()
    assert [row["actual_state"] for row in stored] == [
        "activated",
        "not-activated",
    ]
    assert [row["result"] for row in stored] == ["pass", "fail"]
    assert stored[0]["comments"] == "Witnessed at panel"


def test_zone_geometry_can_be_moved_reassigned_and_unassigned(
    tmp_path: Path,
) -> None:
    project_path = tmp_path / "geometry.fcp"
    sqlite3.connect(project_path).close()
    repository = ProjectRepository(project_path)
    floor_id = repository.add_floor("Ground", 0)
    replacement_dxf = tmp_path / "replacement.dxf"
    replacement_dxf.write_text("placeholder", encoding="utf-8")
    repository.set_floor_dxf(floor_id, replacement_dxf)
    assert Path(repository.fetch_floors()[0]["dxf_path"]) == replacement_dxf
    repository.set_floor_dxf(floor_id, None)
    assert repository.fetch_floors()[0]["dxf_path"] is None
    with repository.connection() as connection:
        connection.executemany(
            "INSERT INTO zones(number, description) VALUES (?, ?)",
            [(10, "Ground floor"), (11, "First floor")],
        )

    original = [(0, 0), (10, 0), (10, 10), (0, 0)]
    repository.assign_zone_geometry(10, floor_id, original, "USER_DRAWN")
    geometry = repository.fetch_zone_geometry()[0]
    moved = [(5, 4), (15, 4), (15, 14), (5, 4)]
    repository.update_zone_geometry(geometry["id"], moved)
    assert json.loads(
        repository.fetch_zone_geometry()[0]["geometry_json"]
    ) == [[5.0, 4.0], [15.0, 4.0], [15.0, 14.0], [5.0, 4.0]]
    repository.reassign_zone_geometry(geometry["id"], 11)
    reassigned = repository.fetch_zone_geometry()[0]
    assert reassigned["zone"] == 11
    repository.remove_zone_geometry(reassigned["id"])
    assert repository.fetch_zone_geometry() == []


def test_zone_can_span_floors_but_has_one_polygon_per_floor(
    tmp_path: Path,
) -> None:
    project_path = tmp_path / "multi-floor-zone.fcp"
    sqlite3.connect(project_path).close()
    repository = ProjectRepository(project_path)
    ground_id = repository.add_floor("Ground", 0)
    first_id = repository.add_floor("First", 1)
    with repository.connection() as connection:
        connection.execute(
            "INSERT INTO zones(number, description) VALUES (10, 'Stair')"
        )

    ground = [(0, 0), (10, 0), (10, 10), (0, 0)]
    first = [(20, 20), (30, 20), (30, 30), (20, 20)]
    repository.assign_zone_geometry(10, ground_id, ground, "USER_DRAWN")
    repository.assign_zone_geometry(10, first_id, first, "USER_DRAWN")

    geometries = repository.fetch_zone_geometry()
    assert [(row["zone"], row["floor_id"]) for row in geometries] == [
        (10, first_id),
        (10, ground_id),
    ]
    assert repository.fetch_zones()[0]["floor_name"] == "First, Ground"

    replacement = [(21, 21), (31, 21), (31, 31), (21, 21)]
    repository.assign_zone_geometry(
        10, first_id, replacement, "REPLACEMENT"
    )
    geometries = repository.fetch_zone_geometry()
    assert len(geometries) == 2
    first_geometry = next(
        row for row in geometries if row["floor_id"] == first_id
    )
    assert json.loads(first_geometry["geometry_json"]) == [
        [21.0, 21.0],
        [31.0, 21.0],
        [31.0, 31.0],
        [21.0, 21.0],
    ]


def test_door_records_both_zones_functions_and_linked_devices(
    tmp_path: Path,
) -> None:
    project_path = tmp_path / "doors.fcp"
    sqlite3.connect(project_path).close()
    repository = ProjectRepository(project_path)
    floor_id = repository.add_floor("Ground", 0)
    with repository.connection() as connection:
        connection.executemany(
            "INSERT INTO zones(number, description) VALUES (?, ?)",
            [(10, "Ward"), (11, "Corridor")],
        )
        snapshot_id = connection.execute(
            """
            INSERT INTO snapshots(
                imported_at, source_name, source_path, sha256,
                versions_json, warnings_json
            ) VALUES ('2026-07-25', 'test', '', 'doors-test', '{}', '[]')
            """
        ).lastrowid
        connection.executemany(
            """
            INSERT INTO devices(
                snapshot_id, stable_key, node, panel, loop, address,
                sub_address, zone, text, product_code, observed_type,
                output_group, output_group_name, ringing_style, record_offset
            ) VALUES (?, ?, 1, 'Panel', 1, ?, 0, ?, ?, 0, ?, ?, ?, '', 0)
            """,
            [
                (
                    snapshot_id,
                    "1/1/20/0",
                    20,
                    10,
                    "Door access relay",
                    "Output relay",
                    20,
                    "Access releases",
                ),
                (
                    snapshot_id,
                    "1/1/21/0",
                    21,
                    11,
                    "Door hold-open magnet",
                    "Output relay",
                    21,
                    "Hold opens",
                ),
            ],
        )

    door_id = repository.create_door(
        "Ward entrance",
        floor_id,
        (10, 20),
        (40, 20),
        10,
        11,
        True,
        "1/1/20/0",
        "LOCKED",
        True,
        "1/1/21/0",
        "HELD OPEN",
        "Release on a fire condition either side.",
        door_type="DOUBLE",
        sprite_position=(25, 20),
        rotation_degrees=90,
    )

    door = repository.fetch_doors(floor_id)[0]
    assert door["id"] == door_id
    assert (door["zone_a"], door["zone_b"]) == (10, 11)
    assert door["access_device_key"] == "1/1/20/0"
    assert door["hold_open_device_key"] == "1/1/21/0"
    assert door["door_type"] == "DOUBLE"
    assert (door["sprite_x"], door["sprite_y"]) == (25, 20)
    assert door["rotation_degrees"] == 90
    assert repository.suggest_door_control_devices(10, 11, "access")[0][
        "stable_key"
    ] == "1/1/20/0"

    repository.update_door(
        door_id,
        "Ward entrance",
        floor_id,
        (10, 20),
        (40, 20),
        10,
        11,
        True,
        "1/1/20/0",
        "UNLOCKED",
        True,
        "1/1/21/0",
        "CLOSED",
        "",
        door_type="SINGLE",
        sprite_position=(30, 25),
        rotation_degrees=45,
    )
    updated = repository.fetch_doors()[0]
    assert updated["access_normal_state"] == "UNLOCKED"
    assert updated["hold_open_normal_state"] == "CLOSED"
    assert updated["door_type"] == "SINGLE"
    assert (updated["sprite_x"], updated["sprite_y"]) == (30, 25)
    assert updated["rotation_degrees"] == 45
    repository.move_door(door_id, 35, 28)
    moved = repository.fetch_doors()[0]
    assert (moved["sprite_x"], moved["sprite_y"]) == (35, 28)
    assert moved["rotation_degrees"] == 45
    assert generate_door_rules(repository) == 4
    door_rules = [
        row
        for row in repository.fetch_rules()
        if row["source"] == "Door drawing suggestion"
    ]
    assert {
        (row["trigger_zone"], row["output_group"], row["action"])
        for row in door_rules
    } == {
        (10, 20, "UNLOCK DOOR"),
        (11, 20, "UNLOCK DOOR"),
        (10, 21, "CLOSE FIRE DOOR"),
        (11, 21, "CLOSE FIRE DOOR"),
    }
    repository.delete_door(door_id)
    assert repository.fetch_doors() == []

    unassigned_id = repository.create_door(
        "Internal unassigned entrance",
        floor_id,
        (0, 0),
        (1, 0),
        10,
        10,
        True,
        None,
        "LOCKED",
        True,
        None,
        "HELD OPEN",
        door_type="SINGLE",
        sprite_position=(50, 50),
    )
    unassigned = repository.fetch_doors()[0]
    assert unassigned["id"] == unassigned_id
    assert (unassigned["zone_a"], unassigned["zone_b"]) == (10, 10)
    assert unassigned["access_device_key"] is None
    assert unassigned["hold_open_device_key"] is None

    repository.update_door(
        unassigned_id,
        "Internal unassigned entrance",
        floor_id,
        (0, 0),
        (1, 0),
        10,
        10,
        True,
        "1/1/20/0",
        "LOCKED",
        True,
        "1/1/21/0",
        "HELD OPEN",
        door_type="SINGLE",
        sprite_position=(50, 50),
    )
    assigned_later = repository.fetch_doors()[0]
    assert assigned_later["access_device_key"] == "1/1/20/0"
    assert assigned_later["hold_open_device_key"] == "1/1/21/0"
    assert generate_door_rules(repository) == 2
    assert all(
        "within zone 10" in row["notes"]
        for row in repository.fetch_rules()
        if row["source"] == "Door drawing suggestion"
    )


def test_existing_door_table_migrates_to_allow_same_zone(tmp_path: Path) -> None:
    project_path = tmp_path / "old-door-check.fcp"
    sqlite3.connect(project_path).close()
    repository = ProjectRepository(project_path)
    with repository.connection() as connection:
        connection.execute("DROP INDEX ix_doors_floor")
        connection.execute("DROP INDEX ix_doors_zones")
        connection.execute("DROP TABLE doors")
        connection.execute(
            """
            CREATE TABLE doors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                floor_id INTEGER NOT NULL REFERENCES floors(id) ON DELETE CASCADE,
                start_x REAL NOT NULL,
                start_y REAL NOT NULL,
                end_x REAL NOT NULL,
                end_y REAL NOT NULL,
                zone_a INTEGER NOT NULL REFERENCES zones(number),
                zone_b INTEGER NOT NULL REFERENCES zones(number),
                has_access_control INTEGER NOT NULL DEFAULT 0,
                access_device_key TEXT,
                access_normal_state TEXT NOT NULL DEFAULT 'LOCKED',
                has_hold_open INTEGER NOT NULL DEFAULT 0,
                hold_open_device_key TEXT,
                hold_open_normal_state TEXT NOT NULL DEFAULT 'HELD OPEN',
                door_type TEXT NOT NULL DEFAULT 'SINGLE',
                sprite_x REAL,
                sprite_y REAL,
                rotation_degrees REAL NOT NULL DEFAULT 0,
                notes TEXT NOT NULL DEFAULT '',
                CHECK(zone_a <> zone_b),
                CHECK(has_access_control = 1 OR has_hold_open = 1)
            )
            """
        )

    repository = ProjectRepository(project_path)
    floor_id = repository.add_floor("Ground", 0)
    with repository.connection() as connection:
        connection.execute(
            "INSERT INTO zones(number, description) VALUES (10, 'Ward')"
        )
        schema = connection.execute(
            "SELECT sql FROM sqlite_master WHERE name = 'doors'"
        ).fetchone()["sql"]
    assert "ZONE_A<>ZONE_B" not in "".join(schema.upper().split())

    repository.create_door(
        "Internal ward door",
        floor_id,
        (0, 0),
        (1, 0),
        10,
        10,
        True,
        None,
        "LOCKED",
        False,
        None,
        "CLOSED",
    )
    assert repository.fetch_doors()[0]["zone_b"] == 10


def test_deleted_dxf_polygon_is_ignored_until_underlay_changes(
    tmp_path: Path,
) -> None:
    project_path = tmp_path / "ignored-polygon.fcp"
    sqlite3.connect(project_path).close()
    repository = ProjectRepository(project_path)
    floor_id = repository.add_floor("Ground", 0)
    points = [(0, 0), (10, 0), (10, 10), (0, 0)]

    repository.ignore_zone_shape(
        floor_id,
        points,
        "ZONE_OUTLINES",
    )
    repository = ProjectRepository(project_path)
    assert repository.fetch_ignored_zone_shape_keys(floor_id) == {
        zone_shape_key(points)
    }

    replacement = tmp_path / "replacement.dxf"
    replacement.write_text("placeholder", encoding="utf-8")
    repository.set_floor_dxf(floor_id, replacement)
    assert repository.fetch_ignored_zone_shape_keys(floor_id) == set()


def test_project_tracks_added_removed_and_modified_matrix_cells(
    tmp_path: Path,
) -> None:
    original_path = _workbook(tmp_path / "matrix.xlsx")
    project_path = tmp_path / "project.fcp"
    sqlite3.connect(project_path).close()
    repository = ProjectRepository(project_path)
    repository.import_cause_effect(original_path)

    revised_path = tmp_path / "matrix-revised.xlsx"
    revised = load_workbook(original_path)
    revised["Cause & Effect"]["F1"] = "RENAMED ZONE SOUNDERS"
    revised["Cause & Effect"]["H1"] = "RENAMED UNUSED SPARE"
    revised["Cause & Effect"]["F7"] = "TA"
    revised["Cause & Effect"]["G7"] = None
    revised["Cause & Effect"]["G8"] = "E"
    revised.save(revised_path)
    revised.close()

    revised_id, _ = repository.import_cause_effect(revised_path)
    changes = repository.fetch_cause_effect_changes(revised_id)

    assert {
        (
            row["stable_key"],
            row["change_type"],
            row["field"],
            row["old_value"],
            row["new_value"],
        )
        for row in changes
    } == {
        (
            "zone 1 -> node 10/group 1",
            "modified",
            "ringing_style",
            "E",
            "TA",
        ),
        (
            "zone 1 -> node 10/group 50",
            "removed",
            None,
            "TA - FIRE DOORS",
            None,
        ),
        (
            "zone 1.11 -> node 10/group 50",
            "added",
            None,
            None,
            "E - FIRE DOORS",
        ),
        (
            "node 10/group 1",
            "modified",
            "output_group_name",
            "ZONE 1 SOUNDERS",
            "RENAMED ZONE SOUNDERS",
        ),
        (
            "node 10/group 99",
            "modified",
            "output_group_name",
            "UNUSED SPARE OUTPUT",
            "RENAMED UNUSED SPARE",
        ),
    }
    assert [
        row["entity"]
        for row in repository.fetch_changes()
    ] == [
        "cause_effect_activation",
        "cause_effect_activation",
        "cause_effect_activation",
        "cause_effect_output_group",
        "cause_effect_output_group",
    ]
