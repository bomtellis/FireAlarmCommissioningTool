import json
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from reportlab.platypus import Paragraph, Table

from firepanel.exports import (
    export_cause_effect_comparison_xlsx,
    export_change_pdf,
)


def _write_table(
    archive: zipfile.ZipFile,
    name: str,
    rows: list[dict],
) -> None:
    table_name = Path(name).stem
    archive.writestr(
        name,
        json.dumps(
            {
                "FDBS": {
                    "Version": 16,
                    "Manager": {
                        "TableList": [
                            {
                                "Name": table_name,
                                "RowList": [
                                    {"RowID": index, "Original": row}
                                    for index, row in enumerate(rows)
                                ],
                            }
                        ]
                    },
                }
            }
        ),
    )


def _configuration(path: Path) -> None:
    with zipfile.ZipFile(path, "w") as archive:
        _write_table(
            archive,
            "tblNode.json",
            [{"NetworkAddress": 7, "NodeName": "NHP Cabin"}],
        )


        _write_table(
            archive,
            "tblOutputGroup.json",
            [
                {"NetworkAddress": 7, "GroupNo": 1, "GroupText": "Sounders"},
                {"NetworkAddress": 7, "GroupNo": 2, "GroupText": "Beacons"},
                {"NetworkAddress": 7, "GroupNo": 3, "GroupText": "Doors"},
            ],
        )
        _write_table(
            archive,
            "tblRingingStyle.json",
            [
                {
                    "NetworkAddress": 7,
                    "StyleNumber": 0,
                    "Description": "Evacuate",
                },
                {
                    "NetworkAddress": 7,
                    "StyleNumber": 2,
                    "Description": "5 Minute Timed Alert",
                },
            ],
        )
        _write_table(
            archive,
            "tblOutputGroupLine.json",
            [
                {
                    "NetworkAddress": 7,
                    "GroupNo": 1,
                    "Operation": 0,
                    "OutputStyleNo": 0,
                    "ZoneFrom": 10,
                    "ZoneTo": 10,
                },
                {
                    "NetworkAddress": 7,
                    "GroupNo": 2,
                    "Operation": 0,
                    "OutputStyleNo": 2,
                    "ZoneFrom": 10,
                    "ZoneTo": 10,
                },
                {
                    "NetworkAddress": 7,
                    "GroupNo": 3,
                    "Operation": 0,
                    "OutputStyleNo": 0,
                    "ZoneFrom": 10,
                    "ZoneTo": 10,
                },
            ],
        )


class _Repository:
    name = "Export test"

    def __init__(self, configuration: Path) -> None:
        self.configuration = configuration

    def fetch_snapshots(self):
        return [
            {
                "source_path": str(self.configuration),
                "source_name": self.configuration.name,
            }
        ]

    def fetch_zones(self):
        return [{"number": 10, "description": "Ground floor"}]

    def fetch_cause_effect_output_groups(self):
        return [
            {
                "target_node": 7,
                "target_node_name": "NHP Cabin",
                "output_group": group,
                "output_group_name": name,
            }
            for group, name in ((1, "Sounders"), (2, "Beacons"), (4, "Access"))
        ]

    def fetch_cause_effect_activations(self):
        return [
            {
                "trigger_zone": "10",
                "trigger_zone_name": "Ground floor",
                "target_node": 7,
                "target_node_name": "NHP Cabin",
                "output_group": group,
                "output_group_name": name,
                "ringing_style": style,
                "reference_status": "Matched",
                "comments": "",
            }
            for group, name, style in (
                (1, "Sounders", "E"),
                (2, "Beacons", "E"),
                (4, "Access", "E"),
            )
        ]

    def latest_cause_effect_import(self):
        return {"source_path": "imported-matrix.xlsx"}

    def fetch_output_groups(self):
        return []


def test_changes_pdf_includes_full_history_and_groups_node_changes(
    tmp_path: Path,
    monkeypatch,
) -> None:
    captured_story = []

    class Document:
        def __init__(self, *_args, **_kwargs):
            pass

        def build(self, story, **_kwargs):
            captured_story.extend(story)

    class Repository:
        name = "Complete history"

        def latest_snapshot_id(self):
            return 2

        def fetch_snapshots(self):
            return [
                {
                    "id": 1,
                    "source_name": "original.skf",
                    "imported_at": "2026-07-25T11:00:00",
                },
                {
                    "id": 2,
                    "source_name": "site.skf",
                    "imported_at": "2026-07-25T12:00:00",
                }
            ]

        def latest_cause_effect_import(self):
            return None

        def fetch_change_details(
            self,
            snapshot_id=None,
            include_all_imports=False,
        ):
            assert snapshot_id is None
            assert include_all_imports
            return [
                {
                    "change_type": "modified",
                    "description": "Device text changed",
                    "stable_key": "2/1/4/0",
                    "node": 2,
                    "panel": "Panel Two",
                    "loop": 1,
                    "address": 4,
                    "sub_address": 0,
                    "zone": 10,
                    "device_text": "Ward optical",
                    "device_type": "Optical detector",
                    "output_group": None,
                    "output_group_name": "",
                    "ringing_style": "",
                    "field": "Device text",
                    "old_value": "Old location",
                    "new_value": "Ward optical",
                    "changed_at": "2026-07-25T12:00:00",
                    "source_name": "site.skf",
                },
                {
                    "change_type": "modified",
                    "description": "Cause & Effect output group changed",
                    "stable_key": "node 7/group 12",
                    "node": None,
                    "panel": "",
                    "loop": None,
                    "address": None,
                    "sub_address": None,
                    "zone": None,
                    "device_text": "",
                    "device_type": "",
                    "output_group": None,
                    "output_group_name": "",
                    "ringing_style": "",
                    "field": "Output group name",
                    "old_value": "Old sounders",
                    "new_value": "Ward sounders",
                    "changed_at": "2026-07-25T12:30:00",
                    "source_name": "matrix.xlsx",
                },
            ]

        def fetch_project_history(self):
            return [
                {
                    "changed_at": "2026-07-25T13:00:00",
                    "change_type": "added",
                    "area": "Output-to-zone assignments",
                    "record_key": '{"node":2,"output_group":12}',
                    "node": 2,
                    "fields": "",
                    "summary": "Output Group: 12; Zone: 10",
                },
                {
                    "changed_at": "2026-07-25T13:05:00",
                    "change_type": "modified",
                    "area": "Zone drawings",
                    "record_key": '{"id":5}',
                    "node": None,
                    "fields": "geometry_json",
                    "summary": "Drawing points changed",
                },
            ]

    monkeypatch.setattr(
        "firepanel.exports.SimpleDocTemplate", Document
    )
    export_change_pdf(Repository(), tmp_path / "changes.pdf")

    headings = [
        item.getPlainText()
        for item in captured_story
        if isinstance(item, Paragraph)
    ]
    assert headings.count("Node 2 - Panel Two") == 1
    assert headings.count("Node 7") == 1
    assert "Total change history" in headings
    assert "Revision 1 - original.skf" in headings
    assert "Revision 2 - site.skf" in headings
    assert "Other project changes" in headings
    assert "Project history" in headings

    tables = [
        item for item in captured_story if isinstance(item, Table)
    ]
    table_text = [
        cell.getPlainText() if isinstance(cell, Paragraph) else str(cell)
        for table in tables
        for row in table._cellvalues
        for cell in row
    ]
    assert "Device text" in table_text
    assert "Output-to-zone assignments" in table_text
    assert "Zone drawings" in table_text

    captured_story.clear()
    export_change_pdf(
        Repository(),
        tmp_path / "selected-changes.pdf",
        revision_ids=[2],
    )
    selected_headings = [
        item.getPlainText()
        for item in captured_story
        if isinstance(item, Paragraph)
    ]
    assert "Revision 1 - original.skf" not in selected_headings
    assert "Revision 2 - site.skf" in selected_headings


def test_comparison_export_has_three_frozen_sheets_and_all_difference_types(
    tmp_path: Path,
) -> None:
    configuration = tmp_path / "site.skf"
    _configuration(configuration)
    destination = tmp_path / "comparison.xlsx"

    export_cause_effect_comparison_xlsx(
        _Repository(configuration),
        destination,
    )

    workbook = load_workbook(destination, data_only=False)
    assert workbook.sheetnames == [
        "NCF-SKF Matrix",
        "Imported Matrix",
        "Differences",
    ]
    assert workbook["NCF-SKF Matrix"].freeze_panes == "C5"
    assert workbook["Imported Matrix"].freeze_panes == "C5"
    assert workbook["Differences"].freeze_panes == "A5"

    differences = workbook["Differences"]
    assert differences["A2"].value == "1 matched links; 3 differences."
    statuses = {
        differences.cell(row=row, column=1).value
        for row in range(5, differences.max_row + 1)
    }
    assert statuses == {
        "Different ringing style",
        "Configuration only",
        "Imported only",
    }
